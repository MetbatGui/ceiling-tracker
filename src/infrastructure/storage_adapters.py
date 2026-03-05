"""저장소 어댑터 구현.

StoragePort를 구현하여 로컬 파일 시스템 및 Google Drive에 데이터를 저장합니다.
"""
import os
import io
from pathlib import Path
from typing import Optional
import pandas as pd
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

from src.domain.ports import StoragePort


class LocalStorageAdapter(StoragePort):
    """로컬 파일 시스템 저장소 Adapter.

    StoragePort를 구현하여 로컬 파일 시스템에 데이터를 저장합니다.

    Attributes:
        base_path (Path): 기본 저장 경로.
    """
    
    def __init__(self, base_path: str = "data"):
        """LocalStorageAdapter 초기화.

        Args:
            base_path (str): 기본 저장 경로 (기본값: "data").
        """
        self.base_path = Path(base_path)
        self.ensure_directory("")  # 기본 경로 생성
        print(f"[LocalStorage] 초기화 완료 (Base: {self.base_path.absolute()})")
    
    def save_dataframe_excel(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        """DataFrame을 Excel 파일로 저장합니다.

        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로 (base_path 상대 경로).
            **kwargs: to_excel 옵션.

        Returns:
            bool: 성공 여부.
        """
        try:
            full_path = self.base_path / path
            self.ensure_directory(str(full_path.parent.relative_to(self.base_path)))
            df.to_excel(full_path, **kwargs)
            print(f"[LocalStorage] [OK] Excel 저장: {path}")
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] Excel 저장 실패 ({path}): {e}")
            return False
    
    def save_dataframe_csv(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        """DataFrame을 CSV 파일로 저장합니다.

        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로 (base_path 상대 경로).
            **kwargs: to_csv 옵션.

        Returns:
            bool: 성공 여부.
        """
        try:
            full_path = self.base_path / path
            self.ensure_directory(str(full_path.parent.relative_to(self.base_path)))
            df.to_csv(full_path, **kwargs)
            print(f"[LocalStorage] [OK] CSV 저장: {path}")
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] CSV 저장 실패 ({path}): {e}")
            return False
    
    def save_workbook(self, book: openpyxl.Workbook, path: str) -> bool:
        """Openpyxl Workbook을 저장합니다.

        Args:
            book (openpyxl.Workbook): 저장할 Workbook 객체.
            path (str): 저장 경로 (base_path 상대 경로).

        Returns:
            bool: 성공 여부.
        """
        try:
            full_path = self.base_path / path
            self.ensure_directory(str(full_path.parent.relative_to(self.base_path)))
            book.save(full_path)
            print(f"[LocalStorage] [OK] Workbook 저장: {path}")
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] Workbook 저장 실패 ({path}): {e}")
            return False
    
    def load_workbook(self, path: str) -> Optional[openpyxl.Workbook]:
        """Excel Workbook을 로드합니다.

        Args:
            path (str): 로드할 파일 경로 (base_path 상대 경로).

        Returns:
            Optional[openpyxl.Workbook]: 로드된 Workbook 객체, 실패 시 None.
        """
        try:
            full_path = self.base_path / path
            return openpyxl.load_workbook(full_path)
        except FileNotFoundError:
            print(f"[LocalStorage] [Warn] 파일 없음: {path}")
            return None
        except Exception as e:
            print(f"[LocalStorage] [Error] Workbook 로드 실패 ({path}): {e}")
            return None
    
    def path_exists(self, path: str) -> bool:
        """경로가 존재하는지 확인합니다.

        Args:
            path (str): 확인할 경로 (base_path 상대 경로).

        Returns:
            bool: 존재 여부.
        """
        full_path = self.base_path / path
        return full_path.exists()
    
    def ensure_directory(self, path: str) -> bool:
        """디렉토리가 없으면 생성합니다.

        Args:
            path (str): 생성할 디렉토리 경로 (base_path 상대 경로).

        Returns:
            bool: 성공 여부.
        """
        try:
            if path == "":
                # 기본 경로 생성
                self.base_path.mkdir(parents=True, exist_ok=True)
            else:
                full_path = self.base_path / path
                full_path.mkdir(parents=True, exist_ok=True)
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] 디렉토리 생성 실패 ({path}): {e}")
            return False

    def load_dataframe(self, path: str, sheet_name: Optional[str] = None, **kwargs) -> pd.DataFrame:
        """Excel 파일에서 DataFrame을 로드합니다.
        
        Args:
            path (str): 파일 경로.
            sheet_name (Optional[str]): 시트 이름.
            **kwargs: 추가 옵션.
            
        Returns:
            pd.DataFrame: 로드된 DataFrame.
        """
        try:
            full_path = self.base_path / path
            if not full_path.exists():
                return pd.DataFrame()
            
            # sheet_name이 None이면 모든 시트를 dict로 반환하므로, 0(첫 번째 시트)으로 설정
            target_sheet = 0 if sheet_name is None else sheet_name
            return pd.read_excel(full_path, sheet_name=target_sheet, **kwargs)
        except Exception as e:
            print(f"[LocalStorage] [Error] DataFrame 로드 실패 ({path}): {e}")
            return pd.DataFrame()

    def get_file(self, path: str) -> Optional[bytes]:
        """파일의 내용을 바이트로 읽어옵니다.
        
        Args:
            path (str): 파일 경로.
            
        Returns:
            Optional[bytes]: 파일 내용, 실패 시 None.
        """
        try:
            full_path = self.base_path / path
            if not full_path.exists():
                return None
            
            with open(full_path, 'rb') as f:
                return f.read()
        except Exception as e:
            print(f"[LocalStorage] [Error] 파일 읽기 실패 ({path}): {e}")
            return None

    def put_file(self, path: str, data: bytes) -> bool:
        """바이트 데이터를 파일로 저장합니다.
        
        Args:
            path (str): 파일 경로.
            data (bytes): 저장할 데이터.
            
        Returns:
            bool: 저장 성공 여부.
        """
        try:
            full_path = self.base_path / path
            self.ensure_directory(str(full_path.parent.relative_to(self.base_path)))
            
            with open(full_path, 'wb') as f:
                f.write(data)
            print(f"[LocalStorage] [OK] 파일 저장: {path}")
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] 파일 저장 실패 ({path}): {e}")
            return False

    def save_parquet(self, df: pd.DataFrame, path: str) -> bool:
        """DataFrame을 Parquet 파일로 저장합니다.

        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로 (base_path 상대 경로).

        Returns:
            bool: 성공 여부.
        """
        try:
            full_path = self.base_path / path
            self.ensure_directory(str(full_path.parent.relative_to(self.base_path)))
            df.to_parquet(full_path, index=False, engine='pyarrow')
            print(f"[LocalStorage] [OK] Parquet 저장: {path}")
            return True
        except Exception as e:
            print(f"[LocalStorage] [Error] Parquet 저장 실패 ({path}): {e}")
            return False

    def load_parquet(self, path: str) -> pd.DataFrame:
        """Parquet 파일에서 DataFrame을 로드합니다.

        Args:
            path (str): 로드할 파일 경로 (base_path 상대 경로).

        Returns:
            pd.DataFrame: 로드된 DataFrame. 파일이 없으면 빈 DataFrame 반환.
        """
        try:
            full_path = self.base_path / path
            if not full_path.exists():
                return pd.DataFrame()
            return pd.read_parquet(full_path, engine='pyarrow')
        except Exception as e:
            print(f"[LocalStorage] [Error] Parquet 로드 실패 ({path}): {e}")
            return pd.DataFrame()


class GoogleDriveAdapter(StoragePort):
    """Google Drive 저장소 Adapter.

    StoragePort를 구현하여 Google Drive에 데이터를 저장하고 로드합니다.
    OAuth 2.0 Token을 사용하여 인증합니다.

    Attributes:
        token_file (str): Token JSON 파일 경로.
        client_secret_file (str): Client Secret JSON 파일 경로 (Refresh용, 선택).
        drive_service (Any): Google Drive API 서비스 객체.
        root_folder_id (str): 루트 폴더 ID (없으면 'root').
    """

    SCOPES = ['https://www.googleapis.com/auth/drive']

    def __init__(
        self, 
        token_file: str, 
        root_folder_name: str = "ceiling-tracker", 
        root_folder_id: Optional[str] = None,
        client_secret_file: Optional[str] = None
    ):
        """GoogleDriveAdapter 초기화.

        Args:
            token_file (str): Token JSON 파일 경로.
            root_folder_name (str): 데이터를 저장할 최상위 폴더 이름 (root_folder_id가 없을 때 사용).
            root_folder_id (Optional[str]): 데이터를 저장할 최상위 폴더 ID (우선순위 높음).
            client_secret_file (Optional[str]): Refresh Token 갱신을 위한 Client Secret 파일 경로.
        
        Raises:
            ValueError: token_file이 제공되지 않은 경우.
            FileNotFoundError: token_file이 존재하지 않는 경우.
        """
        self.token_file = token_file
        self.client_secret_file = client_secret_file
        
        if not self.token_file:
            raise ValueError("token_file must be provided.")
            
        if not os.path.exists(self.token_file):
             raise FileNotFoundError(f"Token file not found: {self.token_file}")

        self.drive_service = self._authenticate()
        
        if root_folder_id:
            self.root_folder_id = root_folder_id
            print(f"[GoogleDrive] 초기화 완료 (지정된 Root ID: {self.root_folder_id})")
        else:
            self.root_folder_id = self._get_or_create_folder(root_folder_name)
            print(f"[GoogleDrive] 초기화 완료 (Root: {root_folder_name}, ID: {self.root_folder_id})")

    def _authenticate(self):
        """Google Drive API 인증 (OAuth 2.0 Token)."""
        try:
            creds = Credentials.from_authorized_user_file(self.token_file, self.SCOPES)
            
            # 토큰 만료 시 갱신 시도
            if creds and creds.expired and creds.refresh_token:
                print("[GoogleDrive] 토큰 만료, 갱신 시도...")
                creds.refresh(Request())
                
                # 갱신된 토큰 저장
                with open(self.token_file, 'w') as token:
                    token.write(creds.to_json())
                    
            return build('drive', 'v3', credentials=creds)
        except Exception as e:
            raise RuntimeError(f"Google Drive 인증 실패: {e}")

    def _get_or_create_folder(self, folder_name: str, parent_id: str = 'root') -> str:
        """폴더를 찾거나 생성합니다.
        
        Args:
            folder_name (str): 폴더 이름.
            parent_id (str): 부모 폴더 ID (기본: 'root').
            
        Returns:
            str: 폴더 ID.
        """
        query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed = false"
        results = self.drive_service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])

        if files:
            return files[0]['id']
        else:
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [parent_id]
            }
            file = self.drive_service.files().create(body=file_metadata, fields='id').execute()
            print(f"[GoogleDrive] 폴더 생성: {folder_name} (ID: {file.get('id')})")
            return file.get('id')

    def _get_file_id(self, path: str) -> Optional[str]:
        """경로(상대 경로)에 해당하는 파일/폴더의 ID를 찾습니다.
        
        Args:
            path (str): 'folder/subfolder/file.ext' 형태의 경로.
            
        Returns:
            Optional[str]: 파일 ID, 없으면 None.
        """
        parts = path.strip("/").split("/")
        current_parent_id = self.root_folder_id
        
        for i, part in enumerate(parts):
            query = f"name = '{part}' and '{current_parent_id}' in parents and trashed = false"
            results = self.drive_service.files().list(q=query, fields="files(id, mimeType)").execute()
            files = results.get('files', [])
            
            if not files:
                return None
            
            current_parent_id = files[0]['id']
            
        return current_parent_id

    def _ensure_path_directories(self, path: str) -> str:
        """파일 경로의 상위 디렉토리들을 생성하고 마지막 부모 폴더 ID를 반환합니다.
        
        Args:
            path (str): 파일 경로.
            
        Returns:
            str: 마지막 부모 폴더 ID.
        """
        parts = path.strip("/").split("/")
        # 파일명 제외
        dir_parts = parts[:-1]
        
        current_parent_id = self.root_folder_id
        for part in dir_parts:
            current_parent_id = self._get_or_create_folder(part, current_parent_id)
            
        return current_parent_id

    def save_dataframe_excel(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        """DataFrame을 Excel 파일로 저장 (업로드).
        
        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로.
            **kwargs: to_excel 옵션.
            
        Returns:
            bool: 성공 여부.
        """
        try:
            # 메모리에 Excel 파일 생성
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, **kwargs)
            output.seek(0)

            self._upload_file(output, path, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            print(f"[GoogleDrive] [OK] Excel 업로드: {path}")
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] Excel 업로드 실패 ({path}): {e}")
            return False

    def save_dataframe_csv(self, df: pd.DataFrame, path: str, **kwargs) -> bool:
        """DataFrame을 CSV 파일로 저장 (업로드).
        
        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로.
            **kwargs: to_csv 옵션.
            
        Returns:
            bool: 성공 여부.
        """
        try:
            # kwargs에서 encoding 추출 (기본값: cp949)
            encoding = kwargs.pop('encoding', 'cp949')
            
            output_str = io.StringIO()
            df.to_csv(output_str, **kwargs)
            
            # 추출한 encoding으로 bytes 변환
            output_bytes = io.BytesIO(output_str.getvalue().encode(encoding))

            self._upload_file(output_bytes, path, 'text/csv')
            print(f"[GoogleDrive] [OK] CSV 업로드: {path} (encoding: {encoding})")
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] CSV 업로드 실패 ({path}): {e}")
            return False

    def save_workbook(self, book: openpyxl.Workbook, path: str) -> bool:
        """Openpyxl Workbook 저장 (업로드).
        
        Args:
            book (openpyxl.Workbook): 저장할 Workbook.
            path (str): 저장 경로.
            
        Returns:
            bool: 성공 여부.
        """
        try:
            output = io.BytesIO()
            book.save(output)
            output.seek(0)

            self._upload_file(output, path, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            print(f"[GoogleDrive] [OK] Workbook 업로드: {path}")
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] Workbook 업로드 실패 ({path}): {e}")
            return False

    def _upload_file(self, data: io.BytesIO, path: str, mime_type: str):
        """파일 업로드 (생성 또는 업데이트).
        
        Args:
            data (io.BytesIO): 파일 데이터.
            path (str): 파일 경로.
            mime_type (str): MIME 타입.
        """
        filename = os.path.basename(path)
        parent_id = self._ensure_path_directories(path)
        
        # 이미 존재하는지 확인
        query = f"name = '{filename}' and '{parent_id}' in parents and trashed = false"
        results = self.drive_service.files().list(q=query, fields="files(id)").execute()
        files = results.get('files', [])

        media = MediaIoBaseUpload(data, mimetype=mime_type, resumable=True)

        if files:
            # 업데이트
            file_id = files[0]['id']
            self.drive_service.files().update(
                fileId=file_id,
                media_body=media
            ).execute()
        else:
            # 생성
            file_metadata = {
                'name': filename,
                'parents': [parent_id]
            }
            self.drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()

    def load_workbook(self, path: str) -> Optional[openpyxl.Workbook]:
        """Excel Workbook 로드 (다운로드).
        
        Args:
            path (str): 파일 경로.
            
        Returns:
            Optional[openpyxl.Workbook]: 로드된 Workbook, 실패 시 None.
        """
        try:
            file_id = self._get_file_id(path)
            if not file_id:
                print(f"[GoogleDrive] [Warn] 파일 없음: {path}")
                return None

            request = self.drive_service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            fh.seek(0)
            return openpyxl.load_workbook(fh)
        except Exception as e:
            print(f"[GoogleDrive] [Error] Workbook 로드 실패 ({path}): {e}")
            return None

    def path_exists(self, path: str) -> bool:
        """경로 존재 여부 확인.
        
        Args:
            path (str): 확인할 경로.
            
        Returns:
            bool: 존재 여부.
        """
        return self._get_file_id(path) is not None

    def ensure_directory(self, path: str) -> bool:
        """디렉토리 생성.
        
        Args:
            path (str): 생성할 디렉토리 경로.
            
        Returns:
            bool: 성공 여부.
        """
        try:
            self._ensure_path_directories(path + "/dummy") # 부모 디렉토리 생성 로직 재사용
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] 디렉토리 생성 실패 ({path}): {e}")
            return False

    def load_dataframe(self, path: str, sheet_name: Optional[str] = None, **kwargs) -> pd.DataFrame:
        """Excel 파일에서 DataFrame을 로드 (다운로드).
        
        Args:
            path (str): 파일 경로.
            sheet_name (Optional[str]): 시트 이름.
            **kwargs: 추가 옵션.
            
        Returns:
            pd.DataFrame: 로드된 DataFrame.
        """
        try:
            file_id = self._get_file_id(path)
            if not file_id:
                return pd.DataFrame()

            request = self.drive_service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            fh.seek(0)
            # sheet_name이 None이면 모든 시트를 dict로 반환하므로, 0(첫 번째 시트)으로 설정
            target_sheet = 0 if sheet_name is None else sheet_name
            return pd.read_excel(fh, sheet_name=target_sheet, **kwargs)
        except Exception as e:
            print(f"[GoogleDrive] [Error] DataFrame 로드 실패 ({path}): {e}")
            return pd.DataFrame()

    def get_file(self, path: str) -> Optional[bytes]:
        """파일의 내용을 바이트로 읽어옵니다 (다운로드).
        
        Args:
            path (str): 파일 경로.
            
        Returns:
            Optional[bytes]: 파일 내용, 실패 시 None.
        """
        try:
            file_id = self._get_file_id(path)
            if not file_id:
                return None

            request = self.drive_service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            fh.seek(0)
            return fh.read()
        except Exception as e:
            print(f"[GoogleDrive] [Error] 파일 다운로드 실패 ({path}): {e}")
            return None

    def put_file(self, path: str, data: bytes) -> bool:
        """바이트 데이터를 파일로 저장합니다 (업로드).
        
        Args:
            path (str): 파일 경로.
            data (bytes): 데이터.
            
        Returns:
            bool: 성공 여부.
        """
        try:
            # MIME 타입 추론 (간단하게)
            if path.endswith('.xlsx'):
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif path.endswith('.csv'):
                mime_type = 'text/csv'
            else:
                mime_type = 'application/octet-stream'

            output = io.BytesIO(data)
            self._upload_file(output, path, mime_type)
            print(f"[GoogleDrive] [OK] 파일 업로드: {path}")
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] 파일 업로드 실패 ({path}): {e}")
            return False

    def save_parquet(self, df: pd.DataFrame, path: str) -> bool:
        """DataFrame을 Parquet 파일로 저장합니다 (업로드).

        Args:
            df (pd.DataFrame): 저장할 DataFrame.
            path (str): 저장 경로.

        Returns:
            bool: 성공 여부.
        """
        try:
            output = io.BytesIO()
            df.to_parquet(output, index=False, engine='pyarrow')
            output.seek(0)
            self._upload_file(output, path, 'application/octet-stream')
            print(f"[GoogleDrive] [OK] Parquet 업로드: {path}")
            return True
        except Exception as e:
            print(f"[GoogleDrive] [Error] Parquet 업로드 실패 ({path}): {e}")
            return False

    def load_parquet(self, path: str) -> pd.DataFrame:
        """Parquet 파일에서 DataFrame을 로드합니다 (다운로드).

        Args:
            path (str): 파일 경로.

        Returns:
            pd.DataFrame: 로드된 DataFrame. 파일이 없으면 빈 DataFrame 반환.
        """
        try:
            file_id = self._get_file_id(path)
            if not file_id:
                return pd.DataFrame()

            request = self.drive_service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            fh.seek(0)
            return pd.read_parquet(fh, engine='pyarrow')
        except Exception as e:
            print(f"[GoogleDrive] [Error] Parquet 다운로드 실패 ({path}): {e}")
            return pd.DataFrame()
