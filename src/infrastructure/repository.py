import pandas as pd
import io
import os
from datetime import datetime, date, timedelta
from typing import List, Optional, Dict
import openpyxl
from pykrx import stock

from src.domain.ports import CohortRepository, StoragePort
from src.domain.model import CeilingCohort, Stock, TrackedStock


# ---------------------------------------------------------------------------
# Parquet 기반 코호트 저장소 (메인 DB)
# ---------------------------------------------------------------------------

class ParquetCohortRepository(CohortRepository):
    """Parquet 파일 기반 상한가 코호트 저장소 구현체.

    코호트 데이터를 tidy(long) 포맷의 Parquet 파일로 저장하고 불러옵니다.

    스키마:
        cohort_date  : 코호트 날짜 (상한가 발생일)
        stock_name   : 종목명
        stock_code   : 종목코드
        new_high_status : 신고가 상태
        initial_price: 상한가 당일 종가 (기준가)
        price_date   : 추적 날짜
        price        : 해당 날 종가
    """

    PARQUET_PATH = "cohorts.parquet"

    def __init__(self, storage: StoragePort, parquet_path: str = PARQUET_PATH):
        """저장소를 초기화합니다.

        Args:
            storage: 파일 I/O를 담당하는 StoragePort 구현체
            parquet_path: Parquet 파일 경로 (storage 기준 상대 경로)
        """
        self.storage = storage
        self.parquet_path = parquet_path

    # ------------------------------------------------------------------
    # CohortRepository 인터페이스 구현
    # ------------------------------------------------------------------

    def save_cohort(self, cohort: CeilingCohort) -> None:
        """코호트를 Parquet에 저장합니다. 기존 데이터와 병합합니다.

        Args:
            cohort: 저장할 CeilingCohort 객체
        """
        new_df = self._cohort_to_dataframe(cohort)
        if new_df.empty:
            return

        existing_df = self.storage.load_parquet(self.parquet_path)

        if existing_df.empty:
            merged_df = new_df
        else:
            # 같은 (cohort_date, stock_code, price_date) 조합을 새 데이터로 교체
            merged_df = self._merge_dataframes(existing_df, new_df)

        self.storage.save_parquet(merged_df, self.parquet_path)
        print(f"[ParquetRepo] 저장 완료: cohort_date={cohort.cohort_date}, "
              f"종목수={len(cohort.stocks)}")

    def load_recent_cohorts(self, limit_days: int,
                            base_date: Optional[date] = None) -> List[CeilingCohort]:
        """최근 N일 이내 코호트를 불러옵니다.

        Args:
            limit_days: 기준일로부터 며칠 전까지 로드할지
            base_date: 기준 날짜 (기본값: 오늘)

        Returns:
            CeilingCohort 리스트
        """
        if base_date is None:
            base_date = date.today()

        cutoff = base_date - timedelta(days=limit_days)
        return self._load_cohorts_where(
            lambda df: df['cohort_date'] >= pd.Timestamp(cutoff)
        )

    def load_cohorts_in_range(self, start_date: date,
                              end_date: date) -> List[CeilingCohort]:
        """날짜 범위의 코호트를 불러옵니다.

        Args:
            start_date: 시작 날짜 (포함)
            end_date: 종료 날짜 (포함)

        Returns:
            CeilingCohort 리스트
        """
        return self._load_cohorts_where(
            lambda df: (df['cohort_date'] >= pd.Timestamp(start_date)) &
                       (df['cohort_date'] <= pd.Timestamp(end_date))
        )

    # ------------------------------------------------------------------
    # 내부 헬퍼
    # ------------------------------------------------------------------

    def _cohort_to_dataframe(self, cohort: CeilingCohort) -> pd.DataFrame:
        """CeilingCohort를 tidy DataFrame으로 변환합니다."""
        rows = []
        for tracked in cohort.stocks:
            # 당일 가격 (initial_price)
            rows.append({
                'cohort_date': cohort.cohort_date,
                'stock_name': tracked.stock.name,
                'stock_code': tracked.stock.code,
                'new_high_status': tracked.new_high_status or '',
                'initial_price': tracked.initial_price,
                'price_date': cohort.cohort_date,
                'price': tracked.initial_price,
            })
            # 추적 가격 히스토리
            for price_date, price in tracked.price_history.items():
                if price_date == cohort.cohort_date:
                    continue  # 당일은 위에서 이미 추가
                rows.append({
                    'cohort_date': cohort.cohort_date,
                    'stock_name': tracked.stock.name,
                    'stock_code': tracked.stock.code,
                    'new_high_status': tracked.new_high_status or '',
                    'initial_price': tracked.initial_price,
                    'price_date': price_date,
                    'price': price,
                })

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        df['cohort_date'] = pd.to_datetime(df['cohort_date'])
        df['price_date'] = pd.to_datetime(df['price_date'])
        return df

    def _merge_dataframes(self, existing: pd.DataFrame,
                          new: pd.DataFrame) -> pd.DataFrame:
        """기존 DataFrame과 신규 DataFrame을 병합합니다.

        동일한 (cohort_date, stock_code, price_date) 조합은 새 데이터로 교체합니다.
        """
        key_cols = ['cohort_date', 'stock_code', 'price_date']

        # 기존 데이터에서 새 데이터와 겹치는 행 제거
        new_keys = new[key_cols]
        mask = existing.set_index(key_cols).index.isin(
            new_keys.set_index(key_cols).index
        )
        existing_filtered = existing[~mask]

        merged = pd.concat([existing_filtered, new], ignore_index=True)
        merged.sort_values(['cohort_date', 'stock_code', 'price_date'], inplace=True)
        merged.reset_index(drop=True, inplace=True)
        return merged

    def _load_cohorts_where(self, condition_fn) -> List[CeilingCohort]:
        """Parquet을 로드하고 조건 함수로 필터링한 뒤 CeilingCohort 리스트로 변환합니다."""
        df = self.storage.load_parquet(self.parquet_path)
        if df.empty:
            return []

        # 날짜 타입 보장
        df['cohort_date'] = pd.to_datetime(df['cohort_date'])
        df['price_date'] = pd.to_datetime(df['price_date'])

        filtered = df[condition_fn(df)]
        if filtered.empty:
            return []

        return self._dataframe_to_cohorts(filtered)

    def _dataframe_to_cohorts(self, df: pd.DataFrame) -> List[CeilingCohort]:
        """tidy DataFrame을 CeilingCohort 리스트로 복원합니다."""
        cohorts: Dict[date, CeilingCohort] = {}

        for cohort_date_ts, cohort_df in df.groupby('cohort_date'):
            cohort_date = cohort_date_ts.date()
            cohort = CeilingCohort(cohort_date=cohort_date)

            for stock_code, stock_df in cohort_df.groupby('stock_code'):
                first_row = stock_df.iloc[0]
                stock_name = first_row['stock_name']
                new_high_status = first_row['new_high_status']
                initial_price = int(first_row['initial_price'])

                cohort.add_stock(stock_name, stock_code, initial_price, new_high_status)
                tracked = cohort.stocks[-1]

                # 가격 히스토리 복원
                for _, row in stock_df.iterrows():
                    price_date = row['price_date'].date()
                    price = int(row['price'])
                    if price_date != cohort_date:
                        tracked.add_price(price_date, price)

            cohorts[cohort_date] = cohort

        return sorted(cohorts.values(), key=lambda c: c.cohort_date)


# ---------------------------------------------------------------------------
# Excel Export 전용 클래스
# ---------------------------------------------------------------------------

class ExcelExporter:
    """Parquet에서 읽어온 코호트 데이터를 Excel 리포트로 변환하는 Export 전용 클래스.

    데이터 저장과는 무관하며, 시각적 서식이 적용된 엑셀 파일을 생성합니다.
    """

    def export(self, cohorts: List[CeilingCohort],
               end_date: Optional[date] = None) -> openpyxl.Workbook:
        """코호트 리스트를 Workbook으로 변환합니다.

        Args:
            cohorts: 변환할 코호트 리스트
            end_date: 데이터 종료일 (None이면 오늘)

        Returns:
            생성된 openpyxl.Workbook 객체
        """
        from openpyxl import Workbook

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for cohort in sorted(cohorts, key=lambda c: c.cohort_date):
            wb = self._add_cohort_sheet(wb, cohort, end_date)

        return wb

    def _add_cohort_sheet(self, wb: openpyxl.Workbook,
                          cohort: CeilingCohort,
                          end_date: Optional[date]) -> openpyxl.Workbook:
        """Workbook에 코호트 시트를 추가합니다."""
        date_slots = self._calculate_date_slots(cohort, end_date)
        headers = self._create_headers(date_slots)

        stocks_data = cohort.get_stocks_data()
        rows, coloring_map, new_high_coloring = self._create_data_rows(
            stocks_data, cohort, date_slots
        )

        sheet_name = cohort.cohort_date.strftime("%y%m%d")
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        self._write_data_to_sheet(ws, headers, rows)
        self._apply_formatting(ws, coloring_map, new_high_coloring, date_slots)

        return wb

    # ------------------------------------------------------------------
    # 날짜 슬롯 / 헤더
    # ------------------------------------------------------------------

    def _calculate_date_slots(self, cohort: CeilingCohort,
                               end_date: Optional[date]) -> List[Optional[date]]:
        from src.domain.constants import TradingConstants

        if end_date is None:
            end_date = date.today()

        valid_trading_days = self._get_trading_days(cohort.cohort_date, end_date)
        date_slots = valid_trading_days[:TradingConstants.FIXED_DATE_SLOTS]
        while len(date_slots) < TradingConstants.FIXED_DATE_SLOTS:
            date_slots.append(None)
        return date_slots

    def _get_trading_days(self, start_date: date, end_date: date) -> List[date]:
        start_str = start_date.strftime("%Y%m%d")
        end_str = end_date.strftime("%Y%m%d")
        try:
            if start_date > end_date:
                return [start_date]
            trading_days_df = stock.get_index_ohlcv_by_date(start_str, end_str, "1001")
            if not trading_days_df.empty:
                return [t.date() for t in trading_days_df.index]
        except Exception:
            pass
        return [start_date]

    def _create_headers(self, date_slots: List[Optional[date]]) -> List[str]:
        date_headers = [d.strftime("%y%m%d") if d else "" for d in date_slots]
        return ['종목명', '신고가'] + date_headers + ['등락률']

    # ------------------------------------------------------------------
    # 데이터 행 생성
    # ------------------------------------------------------------------

    def _create_data_rows(self, stocks_data: List[Dict],
                          cohort: CeilingCohort,
                          date_slots: List[Optional[date]]) -> tuple:
        rows = []
        coloring_map = {}
        new_high_coloring = {}

        for r_idx, s_data in enumerate(stocks_data):
            excel_row = r_idx + 2

            new_high_status = s_data.get('new_high_status', '')
            if new_high_status:
                new_high_coloring[excel_row] = new_high_status

            row_data, row_coloring = self._create_single_row(
                s_data, cohort, date_slots, excel_row
            )
            rows.append(row_data)
            coloring_map.update(row_coloring)

        return rows, coloring_map, new_high_coloring

    def _create_single_row(self, s_data: Dict, cohort: CeilingCohort,
                           date_slots: List[Optional[date]],
                           excel_row: int) -> tuple:
        full_history = s_data['history'].copy()
        full_history[cohort.cohort_date] = s_data['initial_price']

        local_rate = self._calculate_local_rate(full_history, date_slots, s_data)
        row_coloring = self._calculate_consecutive_coloring(
            full_history, cohort.cohort_date, s_data['initial_price'],
            date_slots, excel_row
        )

        row_values = [s_data['name'], s_data.get('new_high_status', '')]
        for d in date_slots:
            if d and d in full_history:
                row_values.append(full_history[d])
            else:
                row_values.append(None)
        row_values.append(f"{local_rate * 100:.1f}%")

        return row_values, row_coloring

    def _calculate_local_rate(self, full_history: Dict[date, int],
                              date_slots: List[Optional[date]],
                              s_data: Dict) -> float:
        last_valid_date = None
        for d in reversed(date_slots):
            if d and d in full_history:
                last_valid_date = d
                break

        if last_valid_date and s_data['initial_price'] > 0:
            return (full_history[last_valid_date] / s_data['initial_price']) - 1.0
        return s_data['current_rate']

    def _calculate_consecutive_coloring(self, full_history: Dict[date, int],
                                        cohort_date: date, initial_price: int,
                                        date_slots: List[Optional[date]],
                                        excel_row: int) -> Dict:
        from src.domain.constants import TradingConstants

        coloring = {}
        sorted_history_dates = sorted(full_history.keys())
        cons_count = 1
        last_price = initial_price

        for d in sorted_history_dates:
            if d <= cohort_date:
                continue
            price = full_history[d]
            if last_price > 0:
                rate = (price - last_price) / last_price
                if rate >= TradingConstants.CEILING_RATE_MIN:
                    cons_count += 1
                    if cons_count >= TradingConstants.CONSECUTIVE_CEILING_MIN and d in date_slots:
                        slot_idx = date_slots.index(d)
                        col_idx = slot_idx + 3
                        coloring[(excel_row, col_idx)] = cons_count
                else:
                    cons_count = 1
            last_price = price

        return coloring

    # ------------------------------------------------------------------
    # 시트 쓰기 / 서식
    # ------------------------------------------------------------------

    def _write_data_to_sheet(self, ws, headers: List[str],
                             rows: List[List]) -> None:
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _apply_formatting(self, ws, coloring_map: Dict,
                          new_high_coloring: Dict,
                          date_slots: List[Optional[date]]) -> None:
        self._apply_consecutive_colors(ws, coloring_map)
        self._apply_new_high_colors(ws, new_high_coloring)
        self._adjust_column_widths(ws)

    def _apply_consecutive_colors(self, ws, coloring_map: Dict) -> None:
        from openpyxl.styles import PatternFill
        from src.domain.constants import TradingConstants, ExcelConstants

        for (r, c), count in coloring_map.items():
            if count >= TradingConstants.CONSECUTIVE_CEILING_MIN:
                color_code = ExcelConstants.CONSECUTIVE_COLORS.get(
                    count, ExcelConstants.CONSECUTIVE_COLOR_DEFAULT
                )
                fill = PatternFill(start_color=color_code, end_color=color_code,
                                   fill_type="solid")
                ws.cell(row=r, column=c).fill = fill

    def _apply_new_high_colors(self, ws, new_high_coloring: Dict) -> None:
        from openpyxl.styles import PatternFill

        new_high_colors = {
            "역·신": "FF0000",
            "역·근": "FFA500",
            "52·신": "FFFF00",
            "52·근": "92d050",
        }
        for row_idx, status in new_high_coloring.items():
            if status in new_high_colors:
                color_code = new_high_colors[status]
                fill = PatternFill(start_color=color_code, end_color=color_code,
                                   fill_type="solid")
                ws.cell(row=row_idx, column=2).fill = fill

    def _adjust_column_widths(self, ws) -> None:
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            col_letter = (
                get_column_letter(column_cells[0].column)
                if isinstance(column_cells[0].column, int)
                else column_cells[0].column_letter
            )
            adjusted_width = (
                (length + 2) * 1.5 if col_letter == 'A'
                else (length + 2) * 1.1
            )
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

        c_width = ws.column_dimensions['C'].width
        for i in range(3, 13):
            ws.column_dimensions[get_column_letter(i)].width = c_width


# ---------------------------------------------------------------------------
# 레거시: 엑셀 기반 코호트 저장소 (마이그레이션 스크립트 전용)
# ---------------------------------------------------------------------------

class ExcelCohortRepository:
    """엑셀 파일 기반 상한가 코호트 저장소 구현체 (레거시, 마이그레이션 전용).

    기존 엑셀 파일에서 코호트 데이터를 읽어서 Parquet으로 이전하는 용도입니다.
    신규 데이터 수집에는 사용하지 않습니다.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load_all_cohorts(self, storage: StoragePort) -> List[CeilingCohort]:
        """엑셀 파일의 모든 코호트를 읽어옵니다."""
        wb = self._load_workbook_safely(storage, self.file_path)
        if wb is None:
            return []

        cohorts = []
        for sheet_name in wb.sheetnames:
            cohort = self._load_cohort_from_sheet(wb, sheet_name)
            if cohort:
                cohorts.append(cohort)
        return cohorts

    def _load_workbook_safely(self, storage, file_path: str):
        if not storage.path_exists(file_path):
            return None
        try:
            return storage.load_workbook(file_path)
        except Exception as e:
            print(f"[ExcelRepo] 엑셀 읽기 실패: {e}")
            return None

    def _load_cohort_from_sheet(self, wb, sheet_name: str) -> Optional[CeilingCohort]:
        cohort_date = self._parse_sheet_date(sheet_name)
        if cohort_date is None:
            return None

        cohort = CeilingCohort(cohort_date=cohort_date)
        ws = wb[sheet_name]
        
        # 헤더 트리밍 (양끝 공백 제거)
        headers = [str(cell.value).strip() if cell.value is not None else None for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            try:
                self._add_stock_from_row(cohort, headers, row, sheet_name)
            except Exception as e:
                print(f"[ExcelRepo] [Warning] 행 로드 실패 ({sheet_name}): {e}")
                continue

        return cohort

    def _parse_sheet_date(self, sheet_name: str) -> Optional[date]:
        try:
            return datetime.strptime(sheet_name, "%y%m%d").date()
        except ValueError:
            return None

    def _add_stock_from_row(self, cohort: CeilingCohort, headers: List,
                            row: tuple, sheet_name: str) -> None:
        import pandas as pd
        row_dict = dict(zip(headers, row))
        name = str(row_dict.get('종목명', '')).strip()
        if not name or pd.isna(name) or name == 'None':
            return

        new_high_status = ''
        if '신고가' in row_dict and row_dict['신고가'] and not pd.isna(row_dict['신고가']):
            new_high_status = str(row_dict['신고가']).strip()

        # 시트 이름과 일치하는 컬럼(당일 가격) 찾기 (공백 제거 후 비교)
        initial_price = 0
        target_col = sheet_name
        
        # row_dict의 키들도 트리밍해서 비교
        trimmed_row_dict = {str(k).strip(): v for k, v in row_dict.items()}
        
        if target_col in trimmed_row_dict and trimmed_row_dict[target_col] is not None:
            try:
                val = str(trimmed_row_dict[target_col]).replace(',', '').strip()
                initial_price = int(float(val))
            except (ValueError, TypeError):
                initial_price = 0

        cohort.add_stock(name, '', initial_price, new_high_status)
        tracked = cohort.stocks[-1]

        for col_name, col_value in trimmed_row_dict.items():
            if col_value is None or pd.isna(col_value):
                continue
            try:
                # '260102' 형식의 날짜 컬럼인지 확인
                col_date = datetime.strptime(str(col_name), "%y%m%d").date()
                val = str(col_value).replace(',', '').strip()
                price = int(float(val))
                if col_date != cohort.cohort_date:
                    tracked.add_price(col_date, price)
            except (ValueError, TypeError):
                continue
