"""도메인 데이터를 Excel Workbook으로 렌더링하는 표현 계층 컴포넌트.

데이터를 저장하지 않으며, 순수하게 시각화만 담당합니다.
외부 의존성 없이 도메인 모델과 거래일 목록만 입력으로 받습니다.
"""
from __future__ import annotations

from datetime import date
from typing import Dict, List, Optional

import openpyxl

from src.domain.model import CeilingCohort
from src.domain.constants import TradingConstants, ExcelConstants


class ExcelRenderer:
    """코호트 데이터를 Excel 리포트로 렌더링하는 Interface 계층 컴포넌트.

    외부에서 trading_days(거래일 목록)와 CeilingCohort 리스트를 전달받아
    openpyxl Workbook 객체를 생성합니다.
    데이터 저장 및 외부 API 호출 책임은 없습니다.
    """

    def render(self,
               cohorts: List[CeilingCohort],
               trading_days: Optional[List[date]] = None,
               end_date: Optional[date] = None) -> openpyxl.Workbook:
        """코호트 리스트를 Workbook으로 렌더링합니다.

        Args:
            cohorts: 렌더링할 코호트 리스트
            trading_days: 표시할 거래일 목록 (외부에서 주입). None이면 빈 슬롯으로 처리.
            end_date: 데이터 종료일 (참고용, 현재는 trading_days 사용)

        Returns:
            openpyxl.Workbook 객체
        """
        from openpyxl import Workbook

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for cohort in sorted(cohorts, key=lambda c: c.cohort_date):
            wb = self._add_cohort_sheet(wb, cohort, trading_days or [], end_date)

        return wb

    # ------------------------------------------------------------------
    # 시트 구성
    # ------------------------------------------------------------------

    def _add_cohort_sheet(self, wb: openpyxl.Workbook,
                          cohort: CeilingCohort,
                          trading_days: List[date],
                          end_date: Optional[date]) -> openpyxl.Workbook:
        """Workbook에 코호트 시트를 추가합니다."""
        date_slots = self._calculate_date_slots(cohort, trading_days)
        headers = self._create_headers(date_slots)

        stocks_data = cohort.get_stocks_data()
        rows, coloring_map, new_high_coloring = self._create_data_rows(
            stocks_data, cohort, date_slots
        )

        sheet_name = cohort.cohort_date.strftime("%y%m%d")
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        self._write_data_to_sheet(ws, headers, rows)
        self._apply_formatting(ws, coloring_map, new_high_coloring, date_slots)

        return wb

    # ------------------------------------------------------------------
    # 날짜 슬롯 / 헤더
    # ------------------------------------------------------------------

    def _calculate_date_slots(self, cohort: CeilingCohort,
                               trading_days: List[date]) -> List[Optional[date]]:
        """코호트 당일 이후의 거래일을 D+0~D+9 슬롯으로 잘라냅니다."""
        cohort_trading_days = [d for d in trading_days if d >= cohort.cohort_date]
        date_slots: List[Optional[date]] = [d for d in cohort_trading_days[:TradingConstants.FIXED_DATE_SLOTS]]
        while len(date_slots) < TradingConstants.FIXED_DATE_SLOTS:
            date_slots.append(None)
        return date_slots

    def _create_headers(self, date_slots: List[Optional[date]]) -> List[str]:
        date_headers = [d.strftime("%y%m%d") if d else "" for d in date_slots]
        return ['종목명', '신고가'] + date_headers + ['등락률']

    # ------------------------------------------------------------------
    # 데이터 행 생성
    # ------------------------------------------------------------------

    def _create_data_rows(self, stocks_data: List[Dict],
                          cohort: CeilingCohort,
                          date_slots: List[Optional[date]]) -> tuple:
        rows = []
        coloring_map = {}
        new_high_coloring = {}

        for r_idx, s_data in enumerate(stocks_data):
            excel_row = r_idx + 2

            new_high_status = s_data.get('new_high_status', '')
            if new_high_status:
                new_high_coloring[excel_row] = new_high_status

            row_data, row_coloring = self._create_single_row(
                s_data, cohort, date_slots, excel_row
            )
            rows.append(row_data)
            coloring_map.update(row_coloring)

        return rows, coloring_map, new_high_coloring

    def _create_single_row(self, s_data: Dict, cohort: CeilingCohort,
                           date_slots: List[Optional[date]],
                           excel_row: int) -> tuple:
        full_history = s_data['history'].copy()
        full_history[cohort.cohort_date] = s_data['initial_price']

        local_rate = self._calculate_local_rate(full_history, date_slots, s_data)
        row_coloring = self._calculate_consecutive_coloring(
            full_history, cohort.cohort_date, s_data['initial_price'],
            date_slots, excel_row
        )

        row_values = [s_data['name'], s_data.get('new_high_status', '')]
        for d in date_slots:
            if d and d in full_history:
                row_values.append(full_history[d])
            else:
                row_values.append(None)
        row_values.append(f"{local_rate * 100:.1f}%")

        return row_values, row_coloring

    def _calculate_local_rate(self, full_history: Dict[date, int],
                              date_slots: List[Optional[date]],
                              s_data: Dict) -> float:
        last_valid_date = None
        for d in reversed(date_slots):
            if d and d in full_history:
                last_valid_date = d
                break

        if last_valid_date and s_data['initial_price'] > 0:
            return (full_history[last_valid_date] / s_data['initial_price']) - 1.0
        return s_data['current_rate']

    def _calculate_consecutive_coloring(self, full_history: Dict[date, int],
                                        cohort_date: date, initial_price: int,
                                        date_slots: List[Optional[date]],
                                        excel_row: int) -> Dict:
        coloring = {}
        sorted_history_dates = sorted(full_history.keys())
        cons_count = 1
        last_price = initial_price

        for d in sorted_history_dates:
            if d <= cohort_date:
                continue
            price = full_history[d]
            if last_price > 0:
                rate = (price - last_price) / last_price
                if rate >= TradingConstants.CEILING_RATE_MIN:
                    cons_count += 1
                    if cons_count >= TradingConstants.CONSECUTIVE_CEILING_MIN and d in date_slots:
                        slot_idx = date_slots.index(d)
                        col_idx = slot_idx + 3
                        coloring[(excel_row, col_idx)] = cons_count
                else:
                    cons_count = 1
            last_price = price

        return coloring

    # ------------------------------------------------------------------
    # 시트 쓰기 / 서식
    # ------------------------------------------------------------------

    def _write_data_to_sheet(self, ws, headers: List[str],
                             rows: List[List]) -> None:
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _apply_formatting(self, ws, coloring_map: Dict,
                          new_high_coloring: Dict,
                          date_slots: List[Optional[date]]) -> None:
        self._apply_consecutive_colors(ws, coloring_map)
        self._apply_new_high_colors(ws, new_high_coloring)
        self._adjust_column_widths(ws)

    def _apply_consecutive_colors(self, ws, coloring_map: Dict) -> None:
        from openpyxl.styles import PatternFill

        for (r, c), count in coloring_map.items():
            if count >= TradingConstants.CONSECUTIVE_CEILING_MIN:
                color_code = ExcelConstants.CONSECUTIVE_COLORS.get(
                    count, ExcelConstants.CONSECUTIVE_COLOR_DEFAULT
                )
                fill = PatternFill(start_color=color_code, end_color=color_code,
                                   fill_type="solid")
                ws.cell(row=r, column=c).fill = fill

    def _apply_new_high_colors(self, ws, new_high_coloring: Dict) -> None:
        from openpyxl.styles import PatternFill

        new_high_colors = {
            "역·신": "FF0000",
            "역·근": "FFA500",
            "52·신": "FFFF00",
            "52·근": "92d050",
        }
        for row_idx, status in new_high_coloring.items():
            if status in new_high_colors:
                color_code = new_high_colors[status]
                fill = PatternFill(start_color=color_code, end_color=color_code,
                                   fill_type="solid")
                ws.cell(row=row_idx, column=2).fill = fill

    def _adjust_column_widths(self, ws) -> None:
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            col_letter = (
                get_column_letter(column_cells[0].column)
                if isinstance(column_cells[0].column, int)
                else column_cells[0].column_letter
            )
            adjusted_width = (
                (length + 2) * 1.5 if col_letter == 'A'
                else (length + 2) * 1.1
            )
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)

        c_width = ws.column_dimensions['C'].width
        for i in range(3, 13):
            ws.column_dimensions[get_column_letter(i)].width = c_width
